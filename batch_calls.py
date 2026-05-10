"""
Batch dispatcher for smart_agent.py.

Reads an Excel sheet of customers and dispatches one outbound call per row,
sequentially. The metadata sent to the agent matches what
smart_agent.entrypoint expects to find in dial_info:
    phone_number, name, amount, debt_date, national_id_last4

Usage:
    # 1. create a template xlsx with two sample rows
    uv run python batch_dispatch.py --init

    # 2. edit customers.xlsx, then dispatch all rows
    uv run python batch_dispatch.py                 # sequential (default)
    uv run python batch_dispatch.py --parallel      # fire all with --gap spacing

    # custom path / agent / wait timeouts
    uv run python batch_dispatch.py --file my_calls.xlsx
    AGENT_NAME=outbound-caller-smart uv run python batch_dispatch.py

In sequential mode (default) the script dispatches one call, polls the
room until the SIP participant has hung up, then dispatches the next.
In parallel mode it fires all dispatches with a configurable gap.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
import uuid
from pathlib import Path

from dotenv import load_dotenv
from livekit import api
from openpyxl import Workbook, load_workbook

load_dotenv(dotenv_path=Path(__file__).parent / ".env")

COLUMNS = ["phone_number", "name", "amount", "debt_date", "national_id_last4"]

DEFAULT_FILE = Path(__file__).parent / "customers.xlsx"


def write_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "customers"
    ws.append(COLUMNS)
    ws.append(["+966555209485", "محمد", "10000", "2023-01-01", "1234"])
    ws.append(["+966555000001", "أحمد", "7500",  "2023-06-15", "5678"])
    for i, _ in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22
    wb.save(path)


def read_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    missing = [c for c in COLUMNS if c not in header]
    if missing:
        raise SystemExit(f"missing columns in {path.name}: {missing}. expected {COLUMNS}")

    idx = {name: header.index(name) for name in COLUMNS}
    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        record = {name: ("" if row[idx[name]] is None else str(row[idx[name]]).strip())
                  for name in COLUMNS}
        if not record["phone_number"]:
            continue
        rows.append(record)
    return rows


async def dispatch_one(lk_api: api.LiveKitAPI, agent_name: str, row: dict[str, str]) -> str:
    room_name = f"outbound-call-{uuid.uuid4().hex[:8]}"
    await lk_api.room.create_room(
        api.CreateRoomRequest(name=room_name, empty_timeout=300)
    )
    metadata = json.dumps({
        "phone_number": row["phone_number"],
        "name": row["name"] or "محمد",
        "amount": row["amount"] or "10000",
        "debt_date": row["debt_date"] or "2023-01-01",
        "national_id_last4": row["national_id_last4"] or "1234",
    }, ensure_ascii=False)
    await lk_api.agent_dispatch.create_dispatch(
        api.CreateAgentDispatchRequest(
            agent_name=agent_name,
            room=room_name,
            metadata=metadata,
        )
    )
    print(f"dispatched {row['phone_number']} ({row['name']}) -> {room_name}")
    return room_name


async def _sip_participant_present(lk_api: api.LiveKitAPI, room_name: str) -> bool:
    """Return True iff the room exists and has a sip-* participant in it."""
    try:
        resp = await lk_api.room.list_participants(
            api.ListParticipantsRequest(room=room_name)
        )
    except Exception:
        # Room was deleted (agent called delete_room on hangup) -> call ended.
        return False
    return any(p.identity.startswith("sip-") for p in resp.participants)


async def wait_for_call_end(
    lk_api: api.LiveKitAPI,
    room_name: str,
    *,
    join_timeout: float = 60.0,
    call_timeout: float = 600.0,
    poll_interval: float = 3.0,
) -> str:
    """Block until the SIP participant has left the room (or never joined).

    Returns one of: 'completed', 'never_connected', 'timeout'.
    """
    deadline_join = asyncio.get_event_loop().time() + join_timeout
    while asyncio.get_event_loop().time() < deadline_join:
        if await _sip_participant_present(lk_api, room_name):
            break
        await asyncio.sleep(poll_interval)
    else:
        return "never_connected"

    deadline_call = asyncio.get_event_loop().time() + call_timeout
    while asyncio.get_event_loop().time() < deadline_call:
        if not await _sip_participant_present(lk_api, room_name):
            return "completed"
        await asyncio.sleep(poll_interval)
    return "timeout"


async def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default=str(DEFAULT_FILE),
                        help="path to xlsx (default: customers.xlsx)")
    parser.add_argument("--init", action="store_true",
                        help="write a template xlsx and exit")
    parser.add_argument("--gap", type=float, default=10.0,
                        help="parallel mode: seconds between dispatches (default: 10)")
    parser.add_argument("--parallel", action="store_true",
                        help="fire all dispatches with --gap spacing (default: sequential)")
    parser.add_argument("--join-timeout", type=float, default=60.0,
                        help="sequential: max seconds to wait for SIP to connect (default: 60)")
    parser.add_argument("--call-timeout", type=float, default=600.0,
                        help="sequential: max seconds for a call before giving up (default: 600)")
    parser.add_argument("--agent",
                        default=os.getenv("AGENT_NAME", "mobily-tawafuq-local"),
                        help="agent_name to dispatch (default: $AGENT_NAME or outbound-caller-smart)")
    args = parser.parse_args()

    path = Path(args.file)

    if args.init:
        if path.exists():
            print(f"refusing to overwrite existing {path}", file=sys.stderr)
            sys.exit(1)
        write_template(path)
        print(f"wrote template: {path}")
        return

    if not path.exists():
        print(f"{path} not found. run with --init to create a template.", file=sys.stderr)
        sys.exit(1)

    rows = read_rows(path)
    if not rows:
        print(f"no rows with phone_number in {path}", file=sys.stderr)
        sys.exit(1)

    mode = "parallel" if args.parallel else "sequential"
    print(f"dispatching {len(rows)} call(s) via agent '{args.agent}' [{mode}]")

    lk_api = api.LiveKitAPI()
    ok = 0
    try:
        for i, row in enumerate(rows):
            try:
                room = await dispatch_one(lk_api, args.agent, row)
                ok += 1
            except Exception as e:
                print(f"FAILED {row['phone_number']}: {e}", file=sys.stderr)
                continue

            if not args.parallel:
                status = await wait_for_call_end(
                    lk_api, room,
                    join_timeout=args.join_timeout,
                    call_timeout=args.call_timeout,
                )
                print(f"  call {row['phone_number']} -> {status}")
            elif i < len(rows) - 1 and args.gap > 0:
                await asyncio.sleep(args.gap)
    finally:
        await lk_api.aclose()

    print(f"done: {ok}/{len(rows)} dispatched")


if __name__ == "__main__":
    asyncio.run(main())